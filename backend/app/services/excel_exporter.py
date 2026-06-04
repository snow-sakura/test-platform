"""Excel 导出服务：将测试用例列表导出为规范格式 Excel"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def export_test_cases_to_excel(cases: list) -> bytes:
    """将测试用例列表导出为 Excel 二进制数据

    Args:
        cases: TestCase 模型对象列表

    Returns:
        Excel 文件字节流
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    # 表头
    headers = ["用例编号", "用例标题", "前置条件", "测试步骤", "预期结果", "优先级", "用例类型"]
    col_widths = [18, 35, 25, 50, 30, 12, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # 数据行
    for row_idx, case in enumerate(cases, 2):
        # 格式化步骤为多行文本
        steps_text = ""
        if case.steps:
            steps_lines = []
            for i, step in enumerate(case.steps, 1):
                if isinstance(step, dict):
                    steps_lines.append(
                        f"步骤{i}: {step.get('step', '')}\n预期: {step.get('expected_result', '')}"
                    )
            steps_text = "\n\n".join(steps_lines)

        row_data = [
            case.case_number or "",
            case.title,
            case.precondition or "",
            steps_text,
            case.expected_result or "",
            case.priority,
            case.case_type or "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment

    # 冻结首行
    ws.freeze_panes = "A2"

    # 输出到字节流
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
