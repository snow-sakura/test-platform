"""Excel 导入/导出工具"""
import io
from typing import Any

import openpyxl
from fastapi import UploadFile
from fastapi.responses import StreamingResponse


def export_to_excel(
    headers: list[str],
    rows: list[list[Any]],
    sheet_name: str = "Sheet1",
    filename: str = "export.xlsx",
) -> StreamingResponse:
    """导出数据到 Excel 并返回 StreamingResponse"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写入表头
    ws.append(headers)

    # 写入数据行
    for row in rows:
        ws.append(row)

    # 写入到 BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def parse_excel(file: UploadFile) -> list[dict]:
    """解析上传的 Excel 文件为 dict 列表（首行为表头）"""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        item = {}
        for i, value in enumerate(row):
            if i < len(headers):
                item[headers[i]] = value
        result.append(item)

    return result
