"""手工测试全生命周期管理 - API 路由"""
from __future__ import annotations

import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from sqlalchemy import case, func, select, cast
from sqlalchemy.types import Date
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.database import get_db
from app.modules.auth.dependencies import get_current_user
from app.modules.rbac.service import require_permission

from .crud import (
    batch_delete_cases, create_attachment, create_case, create_comment,
    create_report, create_report_template, create_review, create_step,
    create_suite, create_version,
    delete_case, delete_comment, delete_report, delete_report_template,
    delete_review, delete_review_template, delete_step, delete_suite,
    delete_version, get_case,
    get_case_comment_count, get_case_step_count, get_cases,
    get_defect_distribution, get_ai_efficiency, get_team_workload,
    get_my_review_tasks, get_plan, get_plans, get_report,
    get_report_template, get_report_templates, get_review,
    get_review_template, get_review_templates,
    get_reviews, get_run, get_runs, get_reports, get_suite, get_suites,
    get_test_management_dashboard_stats, get_version, get_versions,
    create_run, create_review_assignments,
    submit_review_assignment, update_case, update_plan,
    update_report_template,
    update_review, update_review_template, update_run_case, update_step,
    update_suite, update_version,
)
from .models import (
    TestManagementCaseAttachment, TestManagementCaseComment,
    TestManagementCaseStep, TestManagementReviewAssignment,
    TestManagementRunCase,
)
from .schemas import (
    AiEfficiencyItem, DefectDistribution, PlanCreate, PlanResponse, PlanUpdate,
    ReportCreate, ReportResponse,
    ReviewCommentCreate, ReviewCommentResponse,
    ReviewCreate, ReviewDetailResponse, ReviewResponse,
    ReviewTemplateCreate, ReviewTemplateResponse, ReportTemplateResponse,
    ReviewUpdate, ReviewAssignersCreate, ReportTemplateUpdate, RunCaseUpdate, RunResponse,
    TeamWorkloadItem, TestCaseAttachmentResponse, TestCaseCommentCreate,
    TestCaseCommentResponse, TestCaseCreate, TestCaseDetailResponse,
    TestCaseListResponse, TestCaseStepResponse, TestCaseUpdate,
    TestManagementDashboardStats, TestSuiteCreate, TestSuiteDetailResponse,
    TestSuiteResponse, TestSuiteUpdate, VersionCreate, VersionResponse,
    VersionUpdate,
)

router = APIRouter(
    prefix="/test-management",
    dependencies=[Depends(get_current_user)],
    tags=["test-management"],
)


# ==============================
# 仪表盘
# ==============================


@router.get("/dashboard/stats", response_model=TestManagementDashboardStats)
async def dashboard_stats(
    project_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取测试管理仪表盘统计"""
    return await get_test_management_dashboard_stats(db, project_id, current_user.id)


@router.get("/dashboard/execution-trend")
async def execution_trend(
    project_id: int | None = Query(None),
    days: int = Query(7, ge=1, le=90),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """每日执行趋势（默认 7 天）"""
    from datetime import timedelta, datetime, date

    end_date = date.today()
    start_date = end_date - timedelta(days=days - 1)

    from .models import TestManagementRunCase

    query = (
        select(
            cast(TestManagementRunCase.executed_at, Date).label("exec_date"),
            func.count(TestManagementRunCase.id).label("total"),
            func.sum(case((TestManagementRunCase.status == "passed", 1), else_=0)).label("passed"),
            func.sum(case((TestManagementRunCase.status == "failed", 1), else_=0)).label("failed"),
        )
        .where(cast(TestManagementRunCase.executed_at, Date) >= start_date)
        .where(cast(TestManagementRunCase.executed_at, Date) <= end_date)
        .group_by("exec_date")
        .order_by("exec_date")
    )

    result = await db.execute(query)
    rows = result.all()

    date_map = {row.exec_date: {"total": row.total, "passed": row.passed, "failed": row.failed} for row in rows}

    data = []
    for i in range(days):
        d = start_date + timedelta(days=i)
        row_data = date_map.get(d, {"total": 0, "passed": 0, "failed": 0})
        data.append({"date": d.isoformat(), **row_data})

    return {"data": data}


@router.get("/dashboard/status-distribution")
async def status_distribution(
    project_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """执行状态分布"""
    from .models import TestManagementRunCase, TestManagementRun

    query = select(
        TestManagementRunCase.status,
        func.count(TestManagementRunCase.id),
    )

    if project_id:
        query = query.join(TestManagementRun, TestManagementRunCase.run_id == TestManagementRun.id)
        from .models import TestManagementPlan
        query = query.join(TestManagementPlan, TestManagementRun.plan_id == TestManagementPlan.id)
        query = query.where(TestManagementPlan.project_id == project_id)

    query = query.group_by(TestManagementRunCase.status)
    result = await db.execute(query)

    distribution = {row[0]: row[1] for row in result.all()}

    return distribution


@router.get("/dashboard/failed-top10")
async def failed_top10(
    project_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """失败次数最多的 TOP10 用例"""
    from .models import TestManagementRunCase, TestManagementCase, TestManagementRun

    query = (
        select(
            TestManagementCase.id,
            TestManagementCase.title,
            func.count(TestManagementRunCase.id).label("fail_count"),
        )
        .join(TestManagementRunCase, TestManagementRunCase.case_id == TestManagementCase.id)
        .where(TestManagementRunCase.status == "failed")
        .group_by(TestManagementCase.id)
        .order_by(func.count(TestManagementRunCase.id).desc())
        .limit(10)
    )

    if project_id:
        from .models import TestManagementPlan
        query = query.join(TestManagementRun, TestManagementRunCase.run_id == TestManagementRun.id)
        query = query.join(TestManagementPlan, TestManagementRun.plan_id == TestManagementPlan.id)
        query = query.where(TestManagementPlan.project_id == project_id)

    result = await db.execute(query)
    rows = result.all()

    return {"data": [{"case_id": r.id, "title": r.title, "fail_count": r.fail_count} for r in rows]}


@router.get("/dashboard/execution-summary")
async def execution_summary(
    project_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """执行汇总统计"""
    from .models import TestManagementRunCase, TestManagementRun, TestManagementPlan

    base_query = select(TestManagementRunCase.status, func.count().label("cnt"))
    if project_id:
        base_query = base_query.join(TestManagementRun, TestManagementRunCase.run_id == TestManagementRun.id)
        base_query = base_query.join(TestManagementPlan, TestManagementRun.plan_id == TestManagementPlan.id)
        base_query = base_query.where(TestManagementPlan.project_id == project_id)

    total_query = select(func.count()).select_from(base_query.subquery())
    total = (await db.execute(total_query)).scalar() or 0

    # 单次 GROUP BY 查询替代 4 个独立 COUNT
    group_query = base_query.group_by(TestManagementRunCase.status)
    status_rows = (await db.execute(group_query)).all()
    stats = {"total": total, "passed": 0, "failed": 0, "blocked": 0, "untested": 0}
    for row in status_rows:
        if row.status in stats:
            stats[row.status] = row.cnt

    return stats


@router.get("/dashboard/defect-distribution", response_model=DefectDistribution)
async def dashboard_defect_distribution(
    project_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """缺陷分布统计：按用例优先级统计存在缺陷记录的用例数"""
    return await get_defect_distribution(db, project_id)


@router.get("/dashboard/ai-efficiency", response_model=list[AiEfficiencyItem])
async def dashboard_ai_efficiency(
    project_id: int | None = Query(None),
    months: int = Query(12, ge=1, le=36),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """AI 生成 vs 人工创建用例的效能对比（按月聚合）"""
    return await get_ai_efficiency(db, project_id, months)


@router.get("/dashboard/team-workload", response_model=list[TeamWorkloadItem])
async def dashboard_team_workload(
    project_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """团队工作量统计：按执行人统计已执行的用例数"""
    return await get_team_workload(db, project_id)


@router.get("/cases", response_model=dict)
async def list_cases(
    project_id: int = Query(..., description="项目 ID"),
    status: str | None = Query(None),
    priority: str | None = Query(None),
    case_type: str | None = Query(None),
    search: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取项目下的测试用例列表（分页+筛选）"""
    skip = (page - 1) * page_size
    cases, total = await get_cases(db, project_id, status, priority, case_type, search, skip, page_size)

    # 批量计算步骤数和评论数
    case_ids = [c.id for c in cases]
    step_counts = await get_case_step_count(db, case_ids) if case_ids else {}
    comment_counts = await get_case_comment_count(db, case_ids) if case_ids else {}

    items = []
    for c in cases:
        item = TestCaseListResponse.model_validate(c)
        item.step_count = step_counts.get(c.id, 0)
        item.comment_count = comment_counts.get(c.id, 0)
        items.append(item)

    return {"count": total, "results": items}


@router.get("/cases/{case_id}", response_model=TestCaseDetailResponse)
async def retrieve_case(case_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.view"))):
    """获取测试用例详情"""
    case = await get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    return TestCaseDetailResponse.model_validate(case)


@router.post("/cases", response_model=TestCaseDetailResponse, status_code=status.HTTP_201_CREATED)
async def create_new_case(
    project_id: int = Query(..., description="项目 ID"),
    data: TestCaseCreate = ...,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.create")),
):
    """创建测试用例"""
    case = await create_case(db, project_id, current_user.id, data.model_dump())
    case = await get_case(db, case.id)  # 重新加载关联数据
    return TestCaseDetailResponse.model_validate(case)


@router.put("/cases/{case_id}", response_model=TestCaseDetailResponse)
async def update_existing_case(
    case_id: int,
    data: TestCaseUpdate,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.edit")),
):
    """更新测试用例"""
    case = await get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    await update_case(db, case, data.model_dump(exclude_unset=True, mode="json"))
    case = await get_case(db, case_id)
    return TestCaseDetailResponse.model_validate(case)


@router.delete("/cases/{case_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_existing_case(case_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.delete"))):
    """删除测试用例"""
    case = await get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    await delete_case(db, case)


@router.post("/cases/batch-delete", status_code=status.HTTP_204_NO_CONTENT)
async def batch_delete_cases_endpoint(
    ids: list[int],
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.delete")),
):
    """批量删除测试用例"""
    await batch_delete_cases(db, ids)


# ==============================
# Excel 导入/导出
# ==============================


@router.get("/cases/export")
async def export_cases_excel(
    project_id: int = Query(...),
    status: str | None = Query(None),
    priority: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """导出测试用例为 Excel（含步骤）"""
    from app.services.excel_exporter import export_test_cases_excel
    from fastapi.responses import Response

    # 查询当前项目下所有符合条件的用例（最多 10000 条）
    cases, _ = await get_cases(db, project_id, status=status, priority=priority, skip=0, limit=10000)

    try:
        excel_data = export_test_cases_excel(cases)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error("导出用例异常: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="导出失败，请稍后重试")

    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=test_cases_project_{project_id}.xlsx"},
    )


@router.post("/cases/import", status_code=status.HTTP_201_CREATED)
async def import_cases_excel(
    project_id: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.create")),
):
    """从 Excel 导入测试用例（含步骤）"""
    from openpyxl import load_workbook
    import io
    from .crud import batch_import_cases

    # 校验文件格式
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        wb.close()
        raise HTTPException(status_code=400, detail="Excel 文件为空")

    created = 0
    errors = []
    cases_batch = []
    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        try:
            title = str(row[0]).strip()
            if not title:
                continue
            # 截断标题防止过长
            title = title[:500]

            case_data = {
                "title": title,
                "description": str(row[1]) if len(row) > 1 and row[1] else "",
                "preconditions": str(row[2]) if len(row) > 2 and row[2] else "",
                "priority": str(row[3]).upper() if len(row) > 3 and row[3] else "MEDIUM",
                "case_type": str(row[4]) if len(row) > 4 and row[4] else "",
                "steps": [],
            }

            # 解析步骤（第6列为操作，第7列为预期结果，支持多步骤用换行分隔）
            if len(row) > 6 and row[5] and row[6]:
                actions = str(row[5]).split("\n")
                expected = str(row[6]).split("\n")
                for si, act in enumerate(actions):
                    act = act.strip()
                    if act:
                        case_data["steps"].append({
                            "step_number": si + 1,
                            "action": act,
                            "expected_result": expected[si].strip() if si < len(expected) else "",
                        })

            cases_batch.append(case_data)
            created += 1
        except Exception as e:
            errors.append(f"第{i}行: {e}")

    # 批量插入，大幅减少数据库往返
    if cases_batch:
        await batch_import_cases(db, project_id, current_user.id, cases_batch)

    wb.close()
    return {"created": created, "errors": errors, "total": len(rows)}


# ==============================
# 测试步骤
# ==============================


@router.post("/cases/{case_id}/steps", response_model=TestCaseStepResponse, status_code=status.HTTP_201_CREATED)
async def create_new_step(
    case_id: int,
    step_number: int = Query(...),
    action: str = Query(...),
    expected_result: str = Query(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.create")),
):
    """添加测试步骤"""
    case = await get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    step = await create_step(db, case_id, {
        "step_number": step_number, "action": action, "expected_result": expected_result,
    })
    return TestCaseStepResponse.model_validate(step)


@router.put("/cases/{case_id}/steps/{step_id}", response_model=TestCaseStepResponse)
async def update_existing_step(
    case_id: int, step_id: int, data: dict,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.edit")),
):
    """更新测试步骤"""
    step = await db.get(TestManagementCaseStep, step_id)
    if not step or step.case_id != case_id:
        raise HTTPException(status_code=404, detail="步骤不存在")
    step = await update_step(db, step, data)
    return TestCaseStepResponse.model_validate(step)


@router.delete("/cases/{case_id}/steps/{step_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_existing_step(
    case_id: int, step_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.delete")),
):
    """删除测试步骤"""
    step = await db.get(TestManagementCaseStep, step_id)
    if not step or step.case_id != case_id:
        raise HTTPException(status_code=404, detail="步骤不存在")
    await delete_step(db, step)


# ==============================
# 评论
# ==============================


@router.get("/cases/{case_id}/comments", response_model=list[TestCaseCommentResponse])
async def list_case_comments(case_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.view"))):
    """获取用例评论"""
    case = await get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    return [TestCaseCommentResponse.model_validate(c) for c in case.comments]


@router.post("/cases/{case_id}/comments", response_model=TestCaseCommentResponse, status_code=status.HTTP_201_CREATED)
async def create_case_comment(
    case_id: int,
    data: TestCaseCommentCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.create")),
):
    """添加评论"""
    case = await get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    comment = await create_comment(db, case_id, current_user.id, data.content)
    return TestCaseCommentResponse.model_validate(comment)


@router.delete("/cases/{case_id}/comments/{comment_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_case_comment(
    case_id: int, comment_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.delete")),
):
    """删除评论"""
    comment = await db.get(TestManagementCaseComment, comment_id)
    if not comment or comment.case_id != case_id:
        raise HTTPException(status_code=404, detail="评论不存在")
    if comment.author_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能删除自己的评论")
    await delete_comment(db, comment)


# ==============================
# 附件上传
# ==============================


@router.post("/cases/{case_id}/attachments", response_model=TestCaseAttachmentResponse, status_code=status.HTTP_201_CREATED)
async def upload_case_attachment(
    case_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.create")),
):
    """上传测试用例附件"""
    case = await get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="测试用例不存在")

    # 保存文件
    upload_dir = Path(settings.MEDIA_ROOT) / "test_management" / str(case_id)
    upload_dir.mkdir(parents=True, exist_ok=True)

    ext = Path(file.filename).suffix if file.filename else ""
    saved_name = f"{uuid.uuid4().hex}{ext}"
    file_path = upload_dir / saved_name

    content = await file.read()
    file_path.write_bytes(content)

    # 创建记录
    from .crud import create_attachment
    attachment = await create_attachment(
        db, case_id, current_user.id,
        filename=file.filename or saved_name,
        file_path=str(file_path.relative_to(Path(settings.MEDIA_ROOT).parent)),
        file_size=len(content),
    )
    return TestCaseAttachmentResponse.model_validate(attachment)


@router.delete("/cases/{case_id}/attachments/{attachment_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_case_attachment(
    case_id: int, attachment_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.delete")),
):
    """删除附件"""
    att = await db.get(TestManagementCaseAttachment, attachment_id)
    if not att or att.case_id != case_id:
        raise HTTPException(status_code=404, detail="附件不存在")
    # 删除文件
    file_path = Path(settings.MEDIA_ROOT).parent / att.file_path
    if file_path.exists():
        file_path.unlink()
    await db.delete(att)
    await db.flush()


# ==============================
# 测试套件
# ==============================


@router.get("/suites", response_model=dict)
async def list_suites(
    project_id: int = Query(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取项目下的测试套件列表"""
    skip = (page - 1) * page_size
    suites, total = await get_suites(db, project_id, skip, page_size)
    items = []
    for s in suites:
        item = TestSuiteResponse.model_validate(s)
        item.case_count = len(s.case_links) if s.case_links else 0
        items.append(item)
    return {"count": total, "results": items}


@router.get("/suites/{suite_id}", response_model=TestSuiteDetailResponse)
async def retrieve_suite(suite_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.view"))):
    """获取套件详情"""
    suite = await get_suite(db, suite_id)
    if not suite:
        raise HTTPException(status_code=404, detail="套件不存在")
    resp = TestSuiteDetailResponse.model_validate(suite)
    resp.case_count = len(suite.case_links) if suite.case_links else 0
    return resp


@router.post("/suites", response_model=TestSuiteResponse, status_code=status.HTTP_201_CREATED)
async def create_new_suite(
    project_id: int = Query(...),
    data: TestSuiteCreate = ...,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.create")),
):
    """创建测试套件"""
    suite = await create_suite(db, project_id, current_user.id, data.model_dump())
    return TestSuiteResponse.model_validate(suite)


@router.put("/suites/{suite_id}", response_model=TestSuiteResponse)
async def update_existing_suite(suite_id: int, data: TestSuiteUpdate, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.edit"))):
    """更新套件"""
    suite = await get_suite(db, suite_id)
    if not suite:
        raise HTTPException(status_code=404, detail="套件不存在")
    suite = await update_suite(db, suite, data.model_dump(exclude_unset=True, mode="json"))
    return TestSuiteResponse.model_validate(suite)


@router.delete("/suites/{suite_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_existing_suite(suite_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.delete"))):
    """删除套件"""
    suite = await get_suite(db, suite_id)
    if not suite:
        raise HTTPException(status_code=404, detail="套件不存在")
    await delete_suite(db, suite)


# ==============================
# 版本管理
# ==============================


@router.get("/versions", response_model=dict)
async def list_versions(
    project_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取版本列表"""
    skip = (page - 1) * page_size
    versions, total = await get_versions(db, project_id, skip, page_size)
    return {
        "count": total,
        "results": [VersionResponse.model_validate(v) for v in versions],
    }


@router.post("/versions", response_model=VersionResponse, status_code=status.HTTP_201_CREATED)
async def create_new_version(
    data: VersionCreate = ...,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.create")),
):
    """创建版本"""
    version = await create_version(db, current_user.id, data.model_dump())
    return VersionResponse.model_validate(version)


@router.put("/versions/{version_id}", response_model=VersionResponse)
async def update_existing_version(version_id: int, data: VersionUpdate, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.edit"))):
    """更新版本"""
    version = await get_version(db, version_id)
    if not version:
        raise HTTPException(status_code=404, detail="版本不存在")
    version = await update_version(db, version, data.model_dump(exclude_unset=True, mode="json"))
    return VersionResponse.model_validate(version)


@router.delete("/versions/{version_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_existing_version(version_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.delete"))):
    """删除版本"""
    version = await get_version(db, version_id)
    if not version:
        raise HTTPException(status_code=404, detail="版本不存在")
    await delete_version(db, version)


# ==============================
# 评审管理
# ==============================


@router.get("/reviews", response_model=dict)
async def list_reviews(
    project_id: int | None = Query(None),
    status: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取评审列表"""
    skip = (page - 1) * page_size
    reviews, total = await get_reviews(db, project_id, status, skip=skip, limit=page_size)
    return {
        "count": total,
        "results": [ReviewResponse.model_validate(r) for r in reviews],
    }


@router.get("/reviews/my-tasks", response_model=list[ReviewResponse])
async def my_review_tasks(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取我的评审任务"""
    reviews = await get_my_review_tasks(db, current_user.id)
    return [ReviewResponse.model_validate(r) for r in reviews]


@router.get("/reviews/{review_id}", response_model=ReviewDetailResponse)
async def retrieve_review(review_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.view"))):
    """获取评审详情"""
    review = await get_review(db, review_id)
    if not review:
        raise HTTPException(status_code=404, detail="评审不存在")
    resp = ReviewDetailResponse.model_validate(review)
    resp.case_count = len(review.case_links) if review.case_links else 0
    total = len(review.assignments) if review.assignments else 0
    completed = sum(1 for a in review.assignments if a.status == "completed") if review.assignments else 0
    resp.progress = {"total": total, "completed": completed, "pending": total - completed}
    return resp


@router.post("/reviews", response_model=ReviewResponse, status_code=status.HTTP_201_CREATED)
async def create_new_review(
    data: ReviewCreate = ...,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.create")),
):
    """创建评审"""
    review = await create_review(db, current_user.id, data.model_dump())
    return ReviewResponse.model_validate(review)


@router.put("/reviews/{review_id}", response_model=ReviewResponse)
async def update_existing_review(review_id: int, data: ReviewUpdate, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.edit"))):
    """更新评审"""
    review = await get_review(db, review_id)
    if not review:
        raise HTTPException(status_code=404, detail="评审不存在")
    review = await update_review(db, review, data.model_dump(exclude_unset=True, mode="json"))
    return ReviewResponse.model_validate(review)


@router.delete("/reviews/{review_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_existing_review(review_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.delete"))):
    """删除评审"""
    review = await get_review(db, review_id)
    if not review:
        raise HTTPException(status_code=404, detail="评审不存在")
    await delete_review(db, review)


@router.post("/reviews/{review_id}/submit")
async def submit_review(
    review_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.review")),
):
    """提交评审意见"""
    result = await db.execute(
        select(TestManagementReviewAssignment)
        .where(TestManagementReviewAssignment.review_id == review_id)
        .where(TestManagementReviewAssignment.reviewer_id == current_user.id)
    )
    assignment = result.scalar_one_or_none()
    if not assignment:
        raise HTTPException(status_code=404, detail="未找到评审分配记录")
    await submit_review_assignment(db, assignment, data)
    return {"message": "评审已提交"}


@router.post("/reviews/{review_id}/assign-reviewers")
async def assign_reviewers(
    review_id: int,
    data: ReviewAssignersCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.review")),
):
    """为评审分配评审人"""
    review = await get_review(db, review_id)
    if not review:
        raise HTTPException(status_code=404, detail="评审不存在")

    assignments = await create_review_assignments(db, review_id, data.reviewer_ids)

    # 若评审处于草稿状态则更新为进行中
    if review.status == "draft":
        review.status = "in_progress"
        await db.flush()

    return {"message": "评审人已分配", "count": len(assignments)}


# ==============================
# 执行管理
# ==============================


@router.get("/plans", response_model=dict)
async def list_plans(
    project_id: int = Query(...),
    search: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取测试计划列表"""
    skip = (page - 1) * page_size
    plans, total = await get_plans(db, project_id, skip, page_size, search)
    items = []
    for p in plans:
        item = PlanResponse.model_validate(p)
        item.run_count = len(p.runs) if p.runs else 0
        items.append(item)
    return {"count": total, "results": items}


@router.get("/plans/{plan_id}", response_model=PlanResponse)
async def retrieve_plan(plan_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.view"))):
    """获取计划详情"""
    plan = await get_plan(db, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")
    resp = PlanResponse.model_validate(plan)
    resp.run_count = len(plan.runs) if plan.runs else 0
    return resp


@router.post("/plans", response_model=PlanResponse, status_code=status.HTTP_201_CREATED)
async def create_new_plan(
    data: PlanCreate = ...,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.create")),
):
    """创建测试计划"""
    plan = await create_plan(db, current_user.id, data.model_dump())
    return PlanResponse.model_validate(plan)


@router.put("/plans/{plan_id}", response_model=PlanResponse)
async def update_existing_plan(plan_id: int, data: PlanUpdate, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.edit"))):
    """更新计划"""
    plan = await get_plan(db, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")
    await update_plan(db, plan, data.model_dump(exclude_unset=True, mode="json"))
    return PlanResponse.model_validate(plan)


@router.delete("/plans/{plan_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_existing_plan(plan_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.delete"))):
    """删除计划"""
    plan = await get_plan(db, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")
    await delete_plan(db, plan)


@router.post("/plans/{plan_id}/execute")
async def execute_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.execute")),
):
    """执行测试计划：基于计划创建新的执行轮次"""
    plan = await get_plan(db, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")

    # 从计划下所有现有运行的 run_cases 收集用例 ID
    from .models import TestManagementRunCase

    run_ids = [r.id for r in plan.runs]
    if not run_ids:
        raise HTTPException(status_code=400, detail="计划下尚无运行记录，无法执行")

    result = await db.execute(
        select(TestManagementRunCase.case_id)
        .where(TestManagementRunCase.run_id.in_(run_ids))
        .distinct()
    )
    case_ids = [row[0] for row in result.all()]
    if not case_ids:
        raise HTTPException(status_code=400, detail="计划中无测试用例，无法执行")

    run = await create_run(
        db, plan_id,
        name=f"{plan.name} - 第{len(plan.runs) + 1}轮",
        assignee_id=None,
        case_ids=case_ids,
    )
    return {
        "id": run.id,
        "plan_id": run.plan_id,
        "name": run.name,
        "status": run.status,
        "case_count": len(case_ids),
    }


@router.get("/runs", response_model=dict)
async def list_runs(
    plan_id: int | None = Query(None),
    status: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取执行列表"""
    skip = (page - 1) * page_size
    runs, total = await get_runs(db, plan_id, status, skip, page_size)
    items = []
    for r in runs:
        cases = r.run_cases or []
        items.append({
            "id": r.id,
            "plan_id": r.plan_id,
            "name": r.name,
            "assignee_id": r.assignee_id,
            "status": r.status,
            "total_cases": len(cases),
            "passed": sum(1 for c in cases if c.status == "passed"),
            "failed": sum(1 for c in cases if c.status == "failed"),
            "blocked": sum(1 for c in cases if c.status == "blocked"),
            "untested": sum(1 for c in cases if c.status == "untested"),
            "created_at": r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else None,
            "updated_at": r.updated_at.strftime("%Y-%m-%d %H:%M:%S") if r.updated_at else None,
        })
    return {"count": total, "results": items}


@router.get("/runs/{run_id}")
async def retrieve_run(run_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.view"))):
    """获取执行详情（含用例级数据）"""
    run = await get_run(db, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="执行不存在")

    cases = run.run_cases or []
    from .schemas import RunCaseResponse
    return {
        "id": run.id,
        "plan_id": run.plan_id,
        "name": run.name,
        "assignee_id": run.assignee_id,
        "status": run.status,
        "total_cases": len(cases),
        "passed": sum(1 for c in cases if c.status == "passed"),
        "failed": sum(1 for c in cases if c.status == "failed"),
        "blocked": sum(1 for c in cases if c.status == "blocked"),
        "untested": sum(1 for c in cases if c.status == "untested"),
        "run_cases": [RunCaseResponse.model_validate(c).model_dump() for c in cases],
        "created_at": run.created_at.strftime("%Y-%m-%d %H:%M:%S") if run.created_at else None,
        "updated_at": run.updated_at.strftime("%Y-%m-%d %H:%M:%S") if run.updated_at else None,
    }


@router.put("/runs/{run_id}/cases/{run_case_id}")
async def update_run_case_status(
    run_id: int, run_case_id: int,
    data: RunCaseUpdate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.edit")),
):
    """更新执行用例状态"""
    run_case = await db.get(TestManagementRunCase, run_case_id)
    if not run_case or run_case.run_id != run_id:
        raise HTTPException(status_code=404, detail="执行用例不存在")
    await update_run_case(db, run_case, data.model_dump(exclude_unset=True, mode="json"), current_user.id)
    return {"message": "状态已更新"}


# ==============================
# 报告管理
# ==============================


@router.get("/reports", response_model=dict)
async def list_reports(
    project_id: int = Query(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取报告列表"""
    skip = (page - 1) * page_size
    reports, total = await get_reports(db, project_id, skip, page_size)
    return {
        "count": total,
        "results": [ReportResponse.model_validate(r) for r in reports],
    }


@router.post("/reports", response_model=ReportResponse, status_code=status.HTTP_201_CREATED)
async def create_new_report(
    project_id: int = Query(...),
    data: ReportCreate = ...,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission("test_mgmt.create")),
):
    """创建报告"""
    report = await create_report(db, project_id, current_user.id, data.model_dump())
    return ReportResponse.model_validate(report)


@router.delete("/reports/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_existing_report(report_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.delete"))):
    """删除报告"""
    report = await get_report(db, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="报告不存在")
    await delete_report(db, report)


# ====== 评审模板 ======


@router.get("/review-templates", response_model=dict)
async def list_review_templates(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取评审模板列表"""
    skip = (page - 1) * page_size
    templates = await get_review_templates(db, skip, page_size)
    return {"count": len(templates), "results": [ReviewTemplateResponse.model_validate(t) for t in templates]}


@router.get("/review-templates/{template_id}", response_model=ReviewTemplateResponse)
async def retrieve_review_template(template_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.view"))):
    """获取评审模板详情"""
    tmpl = await get_review_template(db, template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="模板不存在")
    return ReviewTemplateResponse.model_validate(tmpl)


@router.post("/review-templates", response_model=ReviewTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_review_template_endpoint(data: ReviewTemplateCreate, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.create"))):
    """创建评审模板"""
    tmpl = await create_review_template(db, data.model_dump())
    return ReviewTemplateResponse.model_validate(tmpl)


@router.put("/review-templates/{template_id}", response_model=ReviewTemplateResponse)
async def update_review_template_endpoint(template_id: int, data: ReviewTemplateCreate, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.edit"))):
    """更新评审模板"""
    tmpl = await get_review_template(db, template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="模板不存在")
    tmpl = await update_review_template(db, tmpl, data.model_dump())
    return ReviewTemplateResponse.model_validate(tmpl)


@router.delete("/review-templates/{template_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_review_template_endpoint(template_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.delete"))):
    """删除评审模板"""
    tmpl = await get_review_template(db, template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="模板不存在")
    await delete_review_template(db, tmpl)


# ====== 报告模板 ======


@router.get("/report-templates", response_model=dict)
async def list_report_templates(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.view")),
):
    """获取报告模板列表"""
    skip = (page - 1) * page_size
    templates = await get_report_templates(db, skip, page_size)
    return {
        "count": len(templates),
        "results": [t.model_dump() if hasattr(t, 'model_dump') else {"id": t.id, "name": t.name} for t in templates],
    }


@router.post("/report-templates", status_code=status.HTTP_201_CREATED)
async def create_report_template_endpoint(data: dict, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.create"))):
    """创建报告模板"""
    tmpl = await create_report_template(db, data)
    return tmpl


@router.get("/report-templates/{template_id}", response_model=ReportTemplateResponse)
async def retrieve_report_template(template_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.view"))):
    """获取报告模板详情"""
    tmpl = await get_report_template(db, template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="报告模板不存在")
    return ReportTemplateResponse.model_validate(tmpl)


@router.put("/report-templates/{template_id}", response_model=ReportTemplateResponse)
async def update_report_template_endpoint(
    template_id: int,
    data: ReportTemplateUpdate,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_permission("test_mgmt.edit")),
):
    """更新报告模板"""
    tmpl = await get_report_template(db, template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="报告模板不存在")
    tmpl = await update_report_template(db, tmpl, data.model_dump(exclude_unset=True, mode="json"))
    return ReportTemplateResponse.model_validate(tmpl)


@router.delete("/report-templates/{template_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_report_template_endpoint(template_id: int, db: AsyncSession = Depends(get_db), _=Depends(require_permission("test_mgmt.delete"))):
    """删除报告模板"""
    tmpl = await get_report_template(db, template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="模板不存在")
    await delete_report_template(db, tmpl)
